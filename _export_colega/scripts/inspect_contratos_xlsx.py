"""Dump del Excel de Codigos de Contratos para entender la estructura."""
import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\Users\Public\Documents\Granos\Codigo de Contratos.xlsx', data_only=True)
for s in wb.worksheets:
    print("="*80)
    print(f"SHEET: {s.title}  ({s.max_row} rows x {s.max_column} cols)")
    print("="*80)
    # Show first 6 rows as raw
    for i, row in enumerate(s.iter_rows(values_only=True)):
        if i >= 8: break
        print(f"row{i}: {row}")
    # Find header row (first row with all non-empty cells)
    print()
    print(f"-- Sample data rows (10 random middle rows) --")
    rows = list(s.iter_rows(values_only=True))
    if len(rows) > 12:
        import math
        for i in range(0, min(len(rows), 16), max(1, len(rows)//8)):
            print(f"r{i}: {rows[i]}")

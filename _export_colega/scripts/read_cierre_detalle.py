"""Lee hojas clave del cierre y muestra contenido completo (truncado)."""
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

XL = Path(r"C:\Users\Public\Documents\Granos\Cierre 29.05.xlsx")
wb = load_workbook(XL, data_only=True, read_only=True)

CLAVE = ["ESTADO PATRIMONIAL", "POSICIÓN GRANARIA", "Silos", "Silo Bolsa", "Bolsas", "Silo Descarte"]

for name in CLAVE:
    if name not in wb.sheetnames:
        print(f"\n--- [{name}] NO ESTÁ ---"); continue
    ws = wb[name]
    print(f"\n{'='*80}\n{'='*5} HOJA: '{name}'  ({ws.max_row}x{ws.max_column}) {'='*5}\n{'='*80}")
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 50: print(f"  ... + {ws.max_row - 50} filas mas"); break
        vals = []
        for v in row:
            if v is None: vals.append("")
            else:
                s = str(v)
                if len(s) > 22: s = s[:22] + "…"
                vals.append(s)
        # quitar columnas vacias del final
        while vals and not vals[-1]: vals.pop()
        if not vals: continue
        print(f"  {i:>3}| " + " | ".join(vals))

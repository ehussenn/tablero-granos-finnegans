"""Dump del Excel de Canje y Proforma para entender la estructura y formulas."""
import sys
sys.stdout.reconfigure(encoding='utf-8')

# .xls usa xlrd, .xlsx usa openpyxl
# Para ver formulas: openpyxl con data_only=False
import openpyxl

# --- Proforma (.xlsx) ---
print("="*100)
print("PROFORMA JULIAI (.xlsx) — con formulas")
print("="*100)
wb = openpyxl.load_workbook(r'C:\Users\Public\Documents\Granos\Proforma Juliai.xlsx', data_only=False)
for s in wb.worksheets:
    print(f"\n>>> SHEET: {s.title}  ({s.max_row} rows x {s.max_column} cols)")
    for row in s.iter_rows(max_row=min(50, s.max_row)):
        cells = []
        for c in row:
            if c.value is None: cells.append("")
            elif isinstance(c.value, str) and c.value.startswith("="):
                cells.append(f"{c.coordinate}={c.value}")
            else:
                cells.append(str(c.value)[:60])
        if any(cells):
            print(" | ".join(cells))

print("\n\n" + "="*100)
print("PROFORMA JULIAI (.xlsx) — con valores")
print("="*100)
wb2 = openpyxl.load_workbook(r'C:\Users\Public\Documents\Granos\Proforma Juliai.xlsx', data_only=True)
for s in wb2.worksheets:
    print(f"\n>>> SHEET: {s.title}")
    for row in s.iter_rows(max_row=min(50, s.max_row), values_only=True):
        if any(c is not None and c != '' for c in row):
            print(" | ".join(str(c) if c is not None else "" for c in row))

# --- Canje (.xls) — necesita xlrd o convertir ---
print("\n\n" + "="*100)
print("CANJE (.xls) — usando xlrd")
print("="*100)
try:
    import xlrd
    wb_xls = xlrd.open_workbook(r'C:\Users\Public\Documents\Granos\Calculo de canje para liquidacion Agronasaja.xls', formatting_info=False)
    for sh in wb_xls.sheets():
        print(f"\n>>> SHEET: {sh.name}  ({sh.nrows} rows x {sh.ncols} cols)")
        for i in range(min(60, sh.nrows)):
            row = sh.row_values(i)
            if any(str(c).strip() for c in row):
                print(" | ".join(str(c)[:60] for c in row))
except ImportError:
    print("xlrd no instalado — instalando...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "xlrd==1.2.0"], check=True)
    import xlrd
    wb_xls = xlrd.open_workbook(r'C:\Users\Public\Documents\Granos\Calculo de canje para liquidacion Agronasaja.xls')
    for sh in wb_xls.sheets():
        print(f"\n>>> SHEET: {sh.name}")
        for i in range(min(60, sh.nrows)):
            row = sh.row_values(i)
            if any(str(c).strip() for c in row):
                print(" | ".join(str(c)[:60] for c in row))

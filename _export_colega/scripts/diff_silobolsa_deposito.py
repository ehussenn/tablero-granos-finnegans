"""Lista depositos de Silo Bolsa que están en Excel pero no captura mi API filter (y viceversa)."""
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
sys.path.insert(0, str(Path(__file__).resolve().parent))
import finnegans_api as api

XL = Path(r"C:\Users\Public\Documents\Granos\Cierre 29.05.xlsx")
wb = load_workbook(XL, data_only=True, read_only=True)
ws = wb["Silo Bolsa"]
header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
i_prod = header.index("Producto"); i_cant = header.index("Stock:cant.1"); i_dep = header.index("Depósito")

# Excel: deposit -> kg (filtrar solo Grano Soja para enfocar)
xl_grano_soja = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or i_prod >= len(row): continue
    prod = row[i_prod]; dep = row[i_dep] if i_dep < len(row) else None
    cant = row[i_cant] if i_cant < len(row) else 0
    try: kg = float(cant or 0)
    except: kg = 0
    if prod == "Grano Soja" and dep:
        xl_grano_soja[dep] = xl_grano_soja.get(dep, 0) + kg

# API: deposit -> kg
api_data = api.call("/reports/USR_RESSTOCKDEP", {"PARAMWEBREPORT_fecha":"getCurrentDate"})
api_grano_soja = {}
for row in api_data:
    if (row.get("PRODUCTO") or "").strip() != "Grano Soja": continue
    dep = row.get("DEPOSITO") or ""
    d = dep.upper()
    es_sb = ("SILOBOLSA" in d) or ("SILO BOLSA" in d)
    if not es_sb: continue
    try: kg = float(row.get("CANTIDAD1") or 0)
    except: kg = 0
    api_grano_soja[dep] = api_grano_soja.get(dep, 0) + kg

# Comparar
xl_total = sum(xl_grano_soja.values()) / 1000
api_total = sum(api_grano_soja.values()) / 1000
print(f"Excel Grano Soja silo bolsa: {xl_total:,.2f} tn ({len(xl_grano_soja)} depositos)")
print(f"API   Grano Soja silo bolsa: {api_total:,.2f} tn ({len(api_grano_soja)} depositos)")
print(f"Diferencia: {xl_total - api_total:,.2f} tn\n")

# Solo Excel (no en mi API)
print("=== Depósitos solo en EXCEL (mi filtro NO los toma) ===")
xl_only = {d: kg for d, kg in xl_grano_soja.items() if d not in api_grano_soja}
for d, kg in sorted(xl_only.items(), key=lambda x: -abs(x[1])):
    if abs(kg) > 0:
        print(f"   {kg/1000:>10,.2f} tn   '{d}'")
sum_xl_only = sum(xl_only.values())/1000
print(f"   SUMA solo-excel: {sum_xl_only:,.2f} tn")

# Solo API (no en Excel)
print("\n=== Depósitos solo en API (Excel NO los muestra) ===")
api_only = {d: kg for d, kg in api_grano_soja.items() if d not in xl_grano_soja}
for d, kg in sorted(api_only.items(), key=lambda x: -abs(x[1]))[:20]:
    if abs(kg) > 0:
        print(f"   {kg/1000:>10,.2f} tn   '{d}'")

# Depositos en AMBOS pero con cantidades distintas
print("\n=== Depósitos en AMBOS con cantidad distinta ===")
for d in sorted(set(xl_grano_soja) & set(api_grano_soja)):
    xv = xl_grano_soja[d]/1000; av = api_grano_soja[d]/1000
    if abs(xv - av) > 0.1:
        print(f"   '{d}': Excel={xv:,.2f}  API={av:,.2f}  diff={xv-av:,.2f}")

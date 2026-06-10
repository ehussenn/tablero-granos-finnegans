"""Compara totales Silo Bolsa de Excel vs API por producto."""
import sys, json, re
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
sys.path.insert(0, str(Path(__file__).resolve().parent))

XL = Path(r"C:\Users\Public\Documents\Granos\Cierre 29.05.xlsx")
wb = load_workbook(XL, data_only=True, read_only=True)
ws = wb["Silo Bolsa"]

# Header columnas - row 2 (1-indexed)
header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
print("HEADER:", header)
# Encuentra indices
i_prod = header.index("Producto") if "Producto" in header else 3
i_cant = header.index("Stock:cant.1") if "Stock:cant.1" in header else 7
i_dep  = header.index("Depósito") if "Depósito" in header else 6
i_conv = header.index("Convertidor") if "Convertidor" in header else 1

print(f"\nIndices: Producto={i_prod}, Cant.1={i_cant}, Depósito={i_dep}, Convertidor={i_conv}")

# Acumular por (Convertidor, Producto)
agg_conv = {}; agg_prod = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row: continue
    prod = row[i_prod] if i_prod < len(row) else None
    conv = row[i_conv] if i_conv < len(row) else None
    cant = row[i_cant] if i_cant < len(row) else 0
    try: kg = float(cant or 0)
    except: kg = 0
    if conv:
        agg_conv[conv] = agg_conv.get(conv, 0) + kg
    if prod:
        agg_prod[prod] = agg_prod.get(prod, 0) + kg

print("\n=== EXCEL Silo Bolsa totales por CONVERTIDOR (usado en POSICIÓN GRANARIA) ===")
for c, kg in sorted(agg_conv.items(), key=lambda x: -abs(x[1])):
    if abs(kg) > 0.01:
        print(f"   {kg/1000:>10,.2f} tn   {c}")

print("\n=== EXCEL Silo Bolsa totales por PRODUCTO ===")
for p, kg in sorted(agg_prod.items(), key=lambda x: -abs(x[1])):
    if abs(kg) > 0.01:
        print(f"   {kg/1000:>10,.2f} tn   {p}")

# Comparar con mi PAYLOAD
ROOT = Path(__file__).resolve().parent.parent
h = (ROOT / "index.html").read_text(encoding="utf-8")
m = re.search(r'const PAYLOAD\s*=\s*(\{.*?\});\s*\n', h, re.S)
P = json.loads(m.group(1))
sb_api = P.get("stock_silobolsa") or {}

print(f"\n=== Comparación: Excel vs API para productos clave ===")
keys = ["Grano Soja", "Grano Maíz", "Grano Trigo Pan", "Grano Arveja", "Grano Girasol"]
print(f"{'Producto':<30} {'Excel (tn)':>12} {'API (tn)':>10} {'Diff':>10}")
for k in keys:
    xl = agg_prod.get(k, 0) / 1000
    api_v = sb_api.get(k, 0)
    print(f"{k:<30} {xl:>12.2f} {api_v:>10.2f} {xl-api_v:>10.2f}")

"""Lee PROYECTADO DE PAGOS GRANOS.xlsx y lo convierte a JSON normalizado."""
import sys, json, re
from pathlib import Path
from datetime import datetime, date
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

SRC = Path(r"C:\Users\Public\Documents\Granos\PROYECTADO DE PAGOS GRANOS.xlsx")
OUT = Path(r"C:\Users\Public\Documents\Granos\tablero-granos-finnegans\data\proyectado_pagos.json")

def to_iso_date(v):
    if v is None or v == "": return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # dd/mm/yyyy o dd-mm-yyyy
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        d, mo, y = m.groups()
        try:
            return date(int(y), int(mo), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            return None
    # yyyy-mm-dd...
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m: return m.group(1)
    return None

def to_float(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    try:
        s = str(v).strip().replace(".", "").replace(",", ".")
        return float(s)
    except ValueError:
        return None

KEY_MAP = {
    "CLIENTE": "cliente",
    "FECHA DE FIJACION": "fecha_fijacion",
    "TN FIJADAS": "tn_fijadas",
    "PRECIO FIJADO": "precio_fijado",
    "TOTAL SIN IVA": "total_sin_iva",
    "TOTAL CON IVA": "total_con_iva",
    "FECHA DE PAGO": "fecha_pago",
}

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]
print(f"Procesando hoja '{ws.title}' ({ws.max_row} filas, {ws.max_column} cols)\n")

# Headers (fila 1, normalizar)
headers = []
for c in range(1, ws.max_column + 1):
    raw = ws.cell(row=1, column=c).value
    if raw is None or str(raw).strip() == "":
        headers.append(None)
        continue
    k = str(raw).strip().upper()
    headers.append(KEY_MAP.get(k, k.lower().replace(" ", "_")))
print(f"Columnas normalizadas: {[h for h in headers if h]}")

rows = []
skipped_empty = 0
for r in range(2, ws.max_row + 1):
    rec = {}
    for c, h in enumerate(headers, start=1):
        if h is None: continue
        rec[h] = ws.cell(row=r, column=c).value

    # Saltear filas totalmente vacias
    if not any(rec.get(k) not in (None, "") for k in rec):
        skipped_empty += 1
        continue

    # Normalizar fechas
    rec["fecha_fijacion"] = to_iso_date(rec.get("fecha_fijacion"))
    rec["fecha_pago"]     = to_iso_date(rec.get("fecha_pago"))
    # Normalizar numericos
    rec["tn_fijadas"]    = to_float(rec.get("tn_fijadas"))
    rec["precio_fijado"] = to_float(rec.get("precio_fijado"))
    rec["total_sin_iva"] = to_float(rec.get("total_sin_iva"))
    rec["total_con_iva"] = to_float(rec.get("total_con_iva"))
    # Cliente con strip
    rec["cliente"] = (rec.get("cliente") or "").strip()
    # Si no tiene cliente y todos los importes/fechas son None, saltear
    if not rec["cliente"] and not rec.get("total_con_iva") and not rec.get("fecha_pago"):
        skipped_empty += 1
        continue

    # Calcular % IVA por fila (si tiene ambos totales)
    if rec.get("total_sin_iva") and rec.get("total_con_iva") and rec["total_sin_iva"] != 0:
        rec["iva_pct"] = round((rec["total_con_iva"] / rec["total_sin_iva"] - 1) * 100, 4)
    else:
        rec["iva_pct"] = None

    # ID único para el frontend
    rec["id"] = f"r{r}"

    rows.append(rec)

print(f"\nTotal filas procesadas: {len(rows)}  (vacías ignoradas: {skipped_empty})")

# Stats
from collections import Counter
total_pagar = sum(r.get("total_con_iva") or 0 for r in rows)
con_fecha = sum(1 for r in rows if r.get("fecha_pago"))
sin_fecha = sum(1 for r in rows if not r.get("fecha_pago"))
print(f"\n📊 Estadísticas:")
print(f"  Total a pagar (suma): USD {total_pagar:,.2f}")
print(f"  Filas con fecha de pago: {con_fecha}")
print(f"  Filas SIN fecha de pago: {sin_fecha}")

# IVA % distintos
iva_dist = Counter(r.get("iva_pct") for r in rows if r.get("iva_pct") is not None)
print(f"\n  % IVA detectados:")
for pct, n in sorted(iva_dist.items(), key=lambda x:-x[1])[:5]:
    print(f"    {pct}% → {n} filas")

# Primer y ultimo registros normalizados
print(f"\nPrimer registro:")
for k, v in rows[0].items():
    print(f"  {k:<20} = {repr(v)[:50]}")
print(f"\nUltimo registro:")
for k, v in rows[-1].items():
    print(f"  {k:<20} = {repr(v)[:50]}")

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\n[+] Guardado: {OUT}  ({OUT.stat().st_size} bytes)")

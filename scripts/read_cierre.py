"""Lee el Excel del cierre 29.05 y muestra las hojas + primeras filas de cada una."""
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

XL = Path(r"C:\Users\Public\Documents\Granos\Cierre 29.05.xlsx")
print(f"Archivo: {XL}  ({XL.stat().st_size:,} bytes)")

wb = load_workbook(XL, data_only=True, read_only=True)
print(f"\nHojas ({len(wb.sheetnames)}):")
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    print(f"  [{i}] '{name}'  →  {rows} filas × {cols} cols")

print("\n" + "="*70)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n--- HOJA: '{name}' ---")
    # primeras 8 filas
    count = 0
    for row in ws.iter_rows(max_row=10, values_only=True):
        # truncar valores largos
        vals = []
        for v in row:
            s = "" if v is None else str(v)
            vals.append(s[:30])
        print("  | " + " | ".join(vals))
        count += 1
    if count == 10:
        print(f"  ... (mas filas)")
